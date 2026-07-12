import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Paths
RAW_CSV_PATH = 'Dataset/customer_support_tickets.csv'
CLEANED_CSV_PATH = 'Dataset/customer_support_tickets_cleaned.csv'
OUTPUT_EXCEL_PATH = 'Excel/customer_support_analysis.xlsx'

def create_excel_workbook():
    print("Initializing Excel Workbook...")
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create sheets
    ws_dashboard = wb.create_sheet(title="KPI Dashboard")
    ws_cleaned = wb.create_sheet(title="Cleaned Data")
    ws_raw = wb.create_sheet(title="Raw Data")
    
    # ----------------------------------------------------
    # STEP 1: Populate Raw Data Sheet
    # ----------------------------------------------------
    print("Populating Raw Data sheet...")
    with open(RAW_CSV_PATH, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for r_idx, row in enumerate(reader, 1):
            for c_idx, val in enumerate(row, 1):
                ws_raw.cell(row=r_idx, column=c_idx, value=val)
                
    # Style Raw Data header
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True)
    for col in range(1, 15):
        cell = ws_raw.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        
    # Enable grid lines
    ws_raw.views.sheetView[0].showGridLines = True
    
    # ----------------------------------------------------
    # STEP 2: Populate Cleaned Data Sheet
    # ----------------------------------------------------
    print("Populating Cleaned Data sheet...")
    with open(CLEANED_CSV_PATH, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for r_idx, row in enumerate(reader, 1):
            for c_idx, val in enumerate(row, 1):
                # Try to cast numeric columns to float/int so Excel treats them as numbers
                # Column 9: Resolution_Time_Hours (Decimal)
                # Column 11: CSAT_Score (Integer)
                cleaned_val = val
                if r_idx > 1: # Skip header
                    if c_idx == 9 and val != "":
                        try:
                            cleaned_val = float(val)
                        except ValueError:
                            pass
                    elif c_idx == 11 and val != "":
                        try:
                            cleaned_val = int(val)
                        except ValueError:
                            pass
                ws_cleaned.cell(row=r_idx, column=c_idx, value=cleaned_val)
                
    # Style Cleaned Data header
    cleaned_header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # light green
    cleaned_header_font = Font(name="Segoe UI", size=10, bold=True, color="375623")
    for col in range(1, 15):
        cell = ws_cleaned.cell(row=1, column=col)
        cell.fill = cleaned_header_fill
        cell.font = cleaned_header_font
        
    # Enable grid lines
    ws_cleaned.views.sheetView[0].showGridLines = True
    
    # Add Conditional Formatting to SLA Column (Column J / 10, cells J2:J2501)
    print("Applying conditional formatting on SLA column...")
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    green_font = Font(color='006100', name="Segoe UI")
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006', name="Segoe UI")
    
    ws_cleaned.conditional_formatting.add('J2:J2501', CellIsRule(operator='equal', formula=['"Met"'], stopIfTrue=True, fill=green_fill, font=green_font))
    ws_cleaned.conditional_formatting.add('J2:J2501', CellIsRule(operator='equal', formula=['"Breached"'], stopIfTrue=True, fill=red_fill, font=red_font))
    
    # ----------------------------------------------------
    # STEP 3: Setup KPI Dashboard Sheet
    # ----------------------------------------------------
    print("Styling KPI Dashboard...")
    ws_dashboard.views.sheetView[0].showGridLines = True
    
    # Common styles
    font_family = "Segoe UI"
    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_label = Font(name=font_family, size=10, bold=True, color="333333")
    font_val = Font(name=font_family, size=18, bold=True, color="1F4E79")
    font_card_lbl = Font(name=font_family, size=9, bold=True, color="FFFFFF")
    
    fill_dark_blue = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_steel_blue = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_light_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_accent_blue = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Title Block
    ws_dashboard.merge_cells('B2:H2')
    title_cell = ws_dashboard['B2']
    title_cell.value = "CUSTOMER SUPPORT INSIGHTS - EXECUTIVE DASHBOARD"
    title_cell.font = font_title
    title_cell.fill = fill_dark_blue
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dashboard.row_dimensions[2].height = 40
    
    # KPI Cards Headers (Row 4)
    kpis = [
        ("B4", "Total Tickets"),
        ("C4", "Closed Tickets"),
        ("D4", "Open Tickets"),
        ("E4", "SLA Compliance %"),
        ("F4", "Avg Resolution (Hrs)"),
        ("G4", "Avg CSAT Score")
    ]
    
    for cell_ref, val in kpis:
        cell = ws_dashboard[cell_ref]
        cell.value = val
        cell.font = font_card_lbl
        cell.fill = fill_steel_blue
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws_dashboard.row_dimensions[4].height = 25
    
    # KPI Formulas (Row 5)
    ws_dashboard['B5'].value = "=COUNTA('Cleaned Data'!A:A)-1"
    ws_dashboard['C5'].value = "=COUNTIF('Cleaned Data'!N:N,\"Closed\")"
    ws_dashboard['D5'].value = "=B5-C5"
    ws_dashboard['E5'].value = "=COUNTIF('Cleaned Data'!J:J,\"Met\")/B5"
    ws_dashboard['F5'].value = "=AVERAGE('Cleaned Data'!I:I)"
    ws_dashboard['G5'].value = "=AVERAGE('Cleaned Data'!K:K)"
    
    for col_char in ["B", "C", "D", "E", "F", "G"]:
        cell = ws_dashboard[f"{col_char}5"]
        cell.font = font_val
        cell.fill = fill_light_blue
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    # Format metrics
    ws_dashboard['B5'].number_format = "#,##0"
    ws_dashboard['C5'].number_format = "#,##0"
    ws_dashboard['D5'].number_format = "#,##0"
    ws_dashboard['E5'].number_format = "0.00%"
    ws_dashboard['F5'].number_format = "0.0"
    ws_dashboard['G5'].number_format = "0.00"
    ws_dashboard.row_dimensions[5].height = 35
    
    # Lookup Tool Header (Row 8)
    ws_dashboard.merge_cells('B8:H8')
    lookup_title = ws_dashboard['B8']
    lookup_title.value = "TICKET QUICK SEARCH TOOL (XLOOKUP ENGINE)"
    lookup_title.font = font_section
    lookup_title.fill = fill_dark_blue
    lookup_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_dashboard.row_dimensions[8].height = 25
    
    # Search input label and field (Row 10)
    ws_dashboard['B10'].value = "Enter Ticket ID:"
    ws_dashboard['B10'].font = font_label
    ws_dashboard['B10'].alignment = Alignment(horizontal="right", vertical="center")
    
    search_cell = ws_dashboard['C10']
    search_cell.value = "TS-1001" # Default value
    search_cell.font = Font(name=font_family, size=11, bold=True, color="1F4E79")
    search_cell.alignment = Alignment(horizontal="center", vertical="center")
    search_cell.border = Border(
        left=Side(style='medium', color='1F4E79'),
        right=Side(style='medium', color='1F4E79'),
        top=Side(style='medium', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79')
    )
    ws_dashboard.row_dimensions[10].height = 25
    
    # Lookup Headers (Row 12)
    lookup_cols = [
        ("B12", "Customer ID"),
        ("C12", "Agent Name"),
        ("D12", "Department"),
        ("E12", "Priority"),
        ("F12", "SLA Status"),
        ("G12", "CSAT Score"),
        ("H12", "Status")
    ]
    
    for cell_ref, val in lookup_cols:
        cell = ws_dashboard[cell_ref]
        cell.value = val
        cell.font = Font(name=font_family, size=10, bold=True, color="000000")
        cell.fill = fill_accent_blue
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    ws_dashboard.row_dimensions[12].height = 22
    
    # Lookup Formulas (Row 13)
    ws_dashboard['B13'].value = "=IF($C$10=\"\",\"\",_xlfn.XLOOKUP($C$10,'Cleaned Data'!$A:$A,'Cleaned Data'!$B:$B,\"Not Found\"))"
    ws_dashboard['C13'].value = "=IF($C$10=\"\",\"\",_xlfn.XLOOKUP($C$10,'Cleaned Data'!$A:$A,'Cleaned Data'!$C:$C,\"Not Found\"))"
    ws_dashboard['D13'].value = "=IF($C$10=\"\",\"\",_xlfn.XLOOKUP($C$10,'Cleaned Data'!$A:$A,'Cleaned Data'!$D:$D,\"Not Found\"))"
    ws_dashboard['E13'].value = "=IF($C$10=\"\",\"\",_xlfn.XLOOKUP($C$10,'Cleaned Data'!$A:$A,'Cleaned Data'!$E:$E,\"Not Found\"))"
    ws_dashboard['F13'].value = "=IF($C$10=\"\",\"\",_xlfn.XLOOKUP($C$10,'Cleaned Data'!$A:$A,'Cleaned Data'!$J:$J,\"Not Found\"))"
    ws_dashboard['G13'].value = "=IF($C$10=\"\",\"\",_xlfn.XLOOKUP($C$10,'Cleaned Data'!$A:$A,'Cleaned Data'!$K:$K,\"Not Found\"))"
    ws_dashboard['H13'].value = "=IF($C$10=\"\",\"\",_xlfn.XLOOKUP($C$10,'Cleaned Data'!$A:$A,'Cleaned Data'!$N:$N,\"Not Found\"))"
    
    for col_char in ["B", "C", "D", "E", "F", "G", "H"]:
        cell = ws_dashboard[f"{col_char}13"]
        cell.font = Font(name=font_family, size=10, color="333333")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    ws_dashboard.row_dimensions[13].height = 25
    
    # ----------------------------------------------------
    # STEP 4: Set Column Widths Auto-fit
    # ----------------------------------------------------
    print("Auto-fitting column widths...")
    # Dashboard Sheet specific adjustments
    ws_dashboard.column_dimensions['A'].width = 3
    ws_dashboard.column_dimensions['B'].width = 22
    ws_dashboard.column_dimensions['C'].width = 22
    ws_dashboard.column_dimensions['D'].width = 20
    ws_dashboard.column_dimensions['E'].width = 22
    ws_dashboard.column_dimensions['F'].width = 22
    ws_dashboard.column_dimensions['G'].width = 22
    ws_dashboard.column_dimensions['H'].width = 20
    
    # Auto-fit Cleaned and Raw Data sheets
    for ws in [ws_cleaned, ws_raw]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    # Save Workbook
    os.makedirs(os.path.dirname(OUTPUT_EXCEL_PATH), exist_ok=True)
    wb.save(OUTPUT_EXCEL_PATH)
    print(f"Excel workbook created successfully at '{OUTPUT_EXCEL_PATH}'.")

if __name__ == '__main__':
    create_excel_workbook()
